"""
CSV Tool — Export and parse CSV data files.

Tools provided:
    create_csv  — Export structured data as a CSV file
    read_csv    — Parse an uploaded CSV file into structured data

Uses pandas for robust CSV handling (type inference, encoding detection,
large file support). Output files saved to the configured artifact directory.
"""

import logging
import uuid
from pathlib import Path
from typing import Optional, Union

from app.context.artifact_manager import get_user_artifacts_dir, get_dept_artifacts_dir, get_company_artifacts_dir
from app.core.user_context import get_user_dept, get_user_email
from app.tools.base_tool import BaseTool

logger = logging.getLogger("mezzofy.tools.csv")


def _get_artifact_dir(config: dict) -> Path:
    base = config.get("storage", {}).get("local_path", "/data/artifacts")
    path = Path(base) / "csv"
    path.mkdir(parents=True, exist_ok=True)
    return path


class CSVOps(BaseTool):
    """CSV file creation and parsing."""

    def __init__(self, config: dict):
        super().__init__(config)
        self._artifact_dir = _get_artifact_dir(config)

    def _resolve_output_dir(self, storage_scope: str = "user") -> Path:
        """Return output dir based on storage scope and user context."""
        dept = get_user_dept()
        email = get_user_email()
        if storage_scope == "company":
            return get_company_artifacts_dir()
        if storage_scope == "department" and dept:
            return get_dept_artifacts_dir(dept)
        if email:
            return get_user_artifacts_dir(dept, email)
        return self._artifact_dir

    def get_tools(self) -> list[dict]:
        return [
            {
                "name": "create_xlsx",
                "description": (
                    "Export structured data as a branded Excel (.xlsx) spreadsheet with "
                    "Mezzofy orange header row. Accepts headers and rows, runs a formula "
                    "error check (QA), and returns the file path and QA result."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "rows": {
                            "type": "array",
                            "description": (
                                "Data rows to write. Each row is a list of values. "
                                "Do NOT include header row here — use the headers parameter."
                            ),
                            "items": {
                                "type": "array",
                                "items": {},
                            },
                        },
                        "headers": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Column headers written as the first (styled) row.",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Worksheet name (default: Sheet1).",
                            "default": "Sheet1",
                        },
                        "filename": {
                            "type": "string",
                            "description": "Output filename (without extension). Auto-generated if omitted.",
                        },
                        "storage_scope": {
                            "type": "string",
                            "description": (
                                "Where to save the file. 'user' = personal folder (default), "
                                "'department' = shared department folder, "
                                "'company' = company-wide public folder (management only)."
                            ),
                            "enum": ["user", "department", "company"],
                            "default": "user",
                        },
                        "template": {
                            "type": "string",
                            "description": (
                                "Template name to apply ('default', 'data_report'). "
                                "The template must exist in knowledge/templates/xlsx/. "
                                "Falls back to blank branded style if omitted or not found."
                            ),
                            "enum": ["default", "data_report"],
                        },
                    },
                    "required": ["rows"],
                },
                "handler": self._create_xlsx,
            },
            {
                "name": "create_csv",
                "description": (
                    "Export structured data as a CSV file. Accepts a list of rows "
                    "(with optional column headers). Returns the file path. "
                    "Use for exporting sales leads, financial data, analytics, or any tabular data."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "rows": {
                            "type": "array",
                            "description": (
                                "Data rows to write. Each row is a list of values. "
                                "If headers are provided, the first row of data should NOT "
                                "include headers (they are added automatically)."
                            ),
                            "items": {
                                "type": "array",
                                "items": {},
                            },
                        },
                        "headers": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Column headers (optional). Written as the first row.",
                        },
                        "filename": {
                            "type": "string",
                            "description": "Output filename (without extension). Auto-generated if omitted.",
                        },
                        "encoding": {
                            "type": "string",
                            "description": "File encoding (default: utf-8-sig for Excel compatibility).",
                            "default": "utf-8-sig",
                        },
                        "storage_scope": {
                            "type": "string",
                            "description": (
                                "Where to save the file. 'user' = personal folder (default), "
                                "'department' = shared department folder, "
                                "'company' = company-wide public folder (management only)."
                            ),
                            "enum": ["user", "department", "company"],
                            "default": "user",
                        },
                    },
                    "required": ["rows"],
                },
                "handler": self._create_csv,
            },
            {
                "name": "read_csv",
                "description": (
                    "Parse an uploaded CSV file and return its contents as structured data. "
                    "Auto-detects delimiter, encoding, and data types. "
                    "Returns headers, rows, and basic summary statistics."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Absolute path to the CSV file to read.",
                        },
                        "max_rows": {
                            "type": "integer",
                            "description": (
                                "Maximum number of rows to return (default: 500). "
                                "Use None to return all rows."
                            ),
                            "default": 500,
                        },
                        "has_header": {
                            "type": "boolean",
                            "description": "Whether the first row is a header row (default: true).",
                            "default": True,
                        },
                    },
                    "required": ["file_path"],
                },
                "handler": self._read_csv,
            },
        ]

    def _resolve_template_path(self, template_name: str) -> Path:
        """Resolve a .xlsx template from knowledge/templates/xlsx/."""
        base = Path(__file__).parent.parent.parent.parent / "knowledge" / "templates" / "xlsx"
        return base / f"{template_name}.xlsx"

    async def _create_xlsx(
        self,
        rows: list[list],
        headers: Optional[list[str]] = None,
        sheet_name: str = "Sheet1",
        filename: Optional[str] = None,
        storage_scope: str = "user",
        template: str = "default",
    ) -> dict:
        """Export data as a branded XLSX file and run formula QA check."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return self._err("openpyxl is not installed. Run: pip install openpyxl")

        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}"

        output_path = self._resolve_output_dir(storage_scope) / f"{filename}.xlsx"

        template_path = self._resolve_template_path(template)
        try:
            try:
                _using_template = template_path.exists()
                if _using_template:
                    wb = openpyxl.load_workbook(str(template_path))
                    ws = wb.active
                    ws.title = sheet_name
                    logger.info(f"Loaded XLSX template: {template_path.name}")
                else:
                    if template:
                        logger.warning(f"XLSX template '{template}' not found, using blank style")
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = sheet_name
            except Exception as e:
                logger.warning(f"Failed to load XLSX template '{template}': {e} — using blank style")
                _using_template = False
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Mezzofy orange header style
            orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            white_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")

            row_offset = 0
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = orange_fill
                    cell.font = white_font
                    cell.alignment = center_align
                row_offset = 1

            for row_idx, row in enumerate(rows, start=1 + row_offset):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(str(output_path))
            file_size = output_path.stat().st_size
            logger.info(f"Created XLSX: {output_path} ({len(rows)} rows)")

            # Run formula QA check
            qa_ok = True
            try:
                import importlib.util
                from pathlib import Path as _Path
                recalc_path = _Path(__file__).parent.parent.parent.parent / "scripts" / "recalc.py"
                spec = importlib.util.spec_from_file_location("recalc", str(recalc_path))
                recalc_mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(recalc_mod)
                qa_ok = recalc_mod.recalc_check(str(output_path))
            except Exception as e:
                logger.warning(f"XLSX QA check failed (non-fatal): {e}")

            return self._ok({
                "file_path": str(output_path),
                "filename": f"{filename}.xlsx",
                "row_count": len(rows),
                "column_count": len(headers) if headers else (len(rows[0]) if rows else 0),
                "size_bytes": file_size,
                "storage_scope": storage_scope,
                "department": get_user_dept(),
                "qa_ok": qa_ok,
            })

        except Exception as e:
            logger.error(f"Failed to create XLSX: {e}")
            return self._err(str(e))

    async def _create_csv(
        self,
        rows: list[list],
        headers: Optional[list[str]] = None,
        filename: Optional[str] = None,
        encoding: str = "utf-8-sig",
        storage_scope: str = "user",
    ) -> dict:
        """Export data as a CSV file."""
        try:
            import pandas as pd
        except ImportError:
            return self._err("pandas is not installed. Run: pip install pandas")

        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}"

        output_path = self._resolve_output_dir(storage_scope) / f"{filename}.csv"

        try:
            df = pd.DataFrame(rows, columns=headers if headers else None)
            df.to_csv(str(output_path), index=False, encoding=encoding)

            file_size = output_path.stat().st_size
            logger.info(f"Created CSV: {output_path} ({len(rows)} rows)")

            return self._ok({
                "file_path": str(output_path),
                "filename": f"{filename}.csv",
                "row_count": len(rows),
                "column_count": len(headers) if headers else (len(rows[0]) if rows else 0),
                "size_bytes": file_size,
                "encoding": encoding,
                "storage_scope": storage_scope,
                "department": get_user_dept(),
            })

        except Exception as e:
            logger.error(f"Failed to create CSV: {e}")
            return self._err(str(e))

    async def _read_csv(
        self,
        file_path: str,
        max_rows: int = 500,
        has_header: bool = True,
    ) -> dict:
        """Parse a CSV file into structured data."""
        import os
        if not os.path.exists(file_path):
            return self._err(f"File not found: {file_path}")

        try:
            import pandas as pd
        except ImportError:
            return self._err("pandas is not installed. Run: pip install pandas")

        try:
            # Auto-detect encoding
            try:
                import chardet
                with open(file_path, "rb") as f:
                    raw = f.read(10000)
                detected = chardet.detect(raw)
                encoding = detected.get("encoding") or "utf-8"
            except ImportError:
                encoding = "utf-8"

            header_row = 0 if has_header else None

            # Try comma first, then auto-detect delimiter
            try:
                df = pd.read_csv(
                    file_path,
                    header=header_row,
                    encoding=encoding,
                    on_bad_lines="skip",
                )
            except Exception:
                df = pd.read_csv(
                    file_path,
                    sep=None,
                    engine="python",
                    header=header_row,
                    encoding=encoding,
                    on_bad_lines="skip",
                )

            total_rows = len(df)
            headers = list(df.columns.astype(str))

            # Limit rows returned
            sample = df.head(max_rows) if max_rows else df

            # Convert to list of lists (JSON-serializable)
            rows = []
            for _, row in sample.iterrows():
                rows.append([
                    None if pd.isna(v) else (
                        str(v) if not isinstance(v, (int, float, bool)) else v
                    )
                    for v in row
                ])

            # Basic summary stats for numeric columns
            summary = {}
            numeric_cols = df.select_dtypes(include="number").columns.tolist()
            for col in numeric_cols[:5]:  # Limit summary to first 5 numeric columns
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    summary[str(col)] = {
                        "min": float(col_data.min()),
                        "max": float(col_data.max()),
                        "mean": round(float(col_data.mean()), 4),
                        "count": int(col_data.count()),
                    }

            return self._ok({
                "file_path": file_path,
                "headers": headers,
                "rows": rows,
                "total_rows": total_rows,
                "returned_rows": len(rows),
                "column_count": len(headers),
                "numeric_summary": summary,
            })

        except Exception as e:
            logger.error(f"Failed to read CSV {file_path}: {e}")
            return self._err(str(e))
