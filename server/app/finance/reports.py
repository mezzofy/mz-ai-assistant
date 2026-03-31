"""Finance report generation — PDF/XLSX using reportlab/openpyxl.

Provides branded report generators for all Mezzofy financial statement types.
All functions gracefully degrade to JSON bytes when optional libraries are
absent (reportlab, openpyxl) so the service never hard-crashes at import time.

Brand colours: Orange #f97316 · Black #000000 · White #ffffff
"""

import logging
from datetime import date
from typing import Optional

logger = logging.getLogger("mezzofy.finance.reports")

# Brand colours (normalised 0-1 for ReportLab)
MEZZOFY_ORANGE = (249 / 255, 115 / 255, 22 / 255)   # #f97316
MEZZOFY_BLACK = (0, 0, 0)
MEZZOFY_WHITE = (1, 1, 1)
MEZZOFY_GRAY = (0.4, 0.4, 0.4)


# ── P&L PDF ───────────────────────────────────────────────────────────────────

async def generate_pnl_pdf(data: dict, entity: dict, period: str) -> bytes:
    """P&L Statement: Revenue → Gross Profit → EBITDA → Net Profit.

    Uses reportlab if available, otherwise returns JSON-encoded bytes.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer,
        )
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        import io

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        story = []
        orange = HexColor("#f97316")

        story.append(Paragraph("<font color='#f97316'><b>MEZZOFY</b></font>", styles["Title"]))
        story.append(Paragraph("Profit &amp; Loss Statement", styles["Heading1"]))
        story.append(Paragraph(
            f"Entity: {entity.get('name', 'N/A')} | Period: {period}", styles["Normal"]
        ))
        story.append(Paragraph(
            f"Generated: {date.today().strftime('%d %b %Y')}", styles["Normal"]
        ))
        story.append(Spacer(1, 10 * mm))

        # Income section
        story.append(Paragraph("<b>INCOME</b>", styles["Heading2"]))
        income_data = [["Category", "Amount (SGD)"]]
        for item in data.get("income", []):
            income_data.append([item["category"], f"{item['amount']:,.2f}"])
        income_data.append([
            "<b>Total Income</b>",
            f"<b>{data.get('total_income', 0):,.2f}</b>",
        ])
        t_income = Table(income_data, colWidths=[120 * mm, 50 * mm])
        t_income.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  orange),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, HexColor("#f9f9f9")]),
            ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEBELOW",     (0, -1), (-1, -1), 1, colors.black),
            ("GRID",          (0, 0), (-1, -1),  0.5, HexColor("#e5e7eb")),
        ]))
        story.append(t_income)
        story.append(Spacer(1, 5 * mm))

        # Expenses section
        story.append(Paragraph("<b>EXPENSES</b>", styles["Heading2"]))
        expense_data = [["Category", "Amount (SGD)"]]
        for item in data.get("expenses", []):
            expense_data.append([item["category"], f"{item['amount']:,.2f}"])
        expense_data.append([
            "<b>Total Expenses</b>",
            f"<b>{data.get('total_expenses', 0):,.2f}</b>",
        ])
        t_expenses = Table(expense_data, colWidths=[120 * mm, 50 * mm])
        t_expenses.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  HexColor("#374151")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, HexColor("#f9f9f9")]),
            ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID",          (0, 0), (-1, -1),  0.5, HexColor("#e5e7eb")),
        ]))
        story.append(t_expenses)
        story.append(Spacer(1, 5 * mm))

        # Net profit
        net = data.get("net_profit", 0)
        net_color = "#16a34a" if net >= 0 else "#dc2626"
        story.append(Paragraph(
            f"<font color='{net_color}'><b>NET PROFIT: SGD {net:,.2f}</b></font>",
            styles["Heading1"],
        ))

        doc.build(story)
        return buffer.getvalue()

    except ImportError:
        logger.warning("reportlab not available — returning JSON bytes for P&L")
        import json
        return json.dumps(data).encode()


# ── AR Aging XLSX ─────────────────────────────────────────────────────────────

async def generate_ar_aging_xlsx(data: dict, entity: dict) -> bytes:
    """AR Aging: Customer | Invoice # | Due Date | Current | 1-30 | 31-60 | 61-90 | 90+"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Aging"

        orange_fill = PatternFill("solid", fgColor="F97316")
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)

        ws["A1"] = f"AR Aging Report — {entity.get('name', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"As at: {data.get('as_at', date.today().strftime('%d %b %Y'))}"
        ws["A3"] = ""

        headers = [
            "Customer / Invoice", "Invoice #", "Due Date",
            "Current", "1-30 days", "31-60 days", "61-90 days", "90+ days",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        buckets = data.get("buckets", {})
        bucket_map = [
            ("current", 4), ("1_30", 5), ("31_60", 6), ("61_90", 7), ("over_90", 8),
        ]
        all_items = []
        for bucket_key, col_idx in bucket_map:
            for item in buckets.get(bucket_key, []):
                item["_bucket_col"] = col_idx
                all_items.append(item)

        row = 5
        for item in sorted(all_items, key=lambda x: x.get("customer", "")):
            ws.cell(row=row, column=1, value=item.get("customer"))
            ws.cell(row=row, column=2, value=item.get("invoice"))
            ws.cell(row=row, column=3, value=item.get("due_date"))
            ws.cell(row=row, column=item["_bucket_col"], value=item.get("amount"))
            row += 1

        total_cell = ws.cell(row=row, column=1, value="TOTAL OUTSTANDING")
        total_cell.font = bold_font
        ws.cell(row=row, column=8, value=data.get("total_outstanding", 0)).font = bold_font

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.warning("openpyxl not available — returning empty bytes for AR aging")
        return b""


# ── Trial Balance XLSX ────────────────────────────────────────────────────────

async def generate_trial_balance_xlsx(data: dict, entity: dict) -> bytes:
    """Trial Balance with debit/credit columns and closing balance."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        orange_fill = PatternFill("solid", fgColor="F97316")
        header_font = Font(bold=True, color="FFFFFF")

        ws["A1"] = f"Trial Balance — {entity.get('name', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"As at: {data.get('as_at', date.today().strftime('%d %b %Y'))}"

        headers = ["Account Code", "Account Name", "Type", "Debit", "Credit", "Balance"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = orange_fill
            cell.font = header_font

        for row_num, acct in enumerate(data.get("accounts", []), 4):
            ws.cell(row=row_num, column=1, value=acct.get("code"))
            ws.cell(row=row_num, column=2, value=acct.get("name"))
            ws.cell(row=row_num, column=3, value=acct.get("type"))
            ws.cell(row=row_num, column=4, value=acct.get("debit"))
            ws.cell(row=row_num, column=5, value=acct.get("credit"))
            ws.cell(row=row_num, column=6, value=acct.get("balance"))

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.warning("openpyxl not available — returning empty bytes for trial balance")
        return b""


# ── Invoice PDF ───────────────────────────────────────────────────────────────

async def generate_invoice_pdf(invoice: dict, entity: dict) -> bytes:
    """Branded customer invoice PDF with line items, tax, and totals."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer,
        )
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        import io

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        story = []
        orange = HexColor("#f97316")

        story.append(Paragraph("<font color='#f97316'><b>MEZZOFY</b></font>", styles["Title"]))
        story.append(Paragraph("<b>INVOICE</b>", styles["Heading1"]))
        story.append(Spacer(1, 5 * mm))

        meta = [
            ["Invoice Number:", invoice.get("invoice_number", "")],
            ["Invoice Date:", str(invoice.get("invoice_date", ""))],
            ["Due Date:", str(invoice.get("due_date", ""))],
            ["Customer:", invoice.get("customer_name", "")],
        ]
        meta_table = Table(meta, colWidths=[50 * mm, 120 * mm])
        story.append(meta_table)
        story.append(Spacer(1, 5 * mm))

        line_items = invoice.get("line_items", [])
        if line_items:
            items_data = [["Description", "Qty", "Unit Price", "Tax %", "Amount"]]
            for item in line_items:
                items_data.append([
                    item.get("description", ""),
                    str(item.get("quantity", "")),
                    f"{float(item.get('unit_price', 0)):,.2f}",
                    f"{float(item.get('tax_rate', 0)):.1f}%",
                    f"{float(item.get('amount', 0)):,.2f}",
                ])
            t = Table(items_data, colWidths=[80 * mm, 20 * mm, 30 * mm, 20 * mm, 30 * mm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0),  orange),
                ("TEXTCOLOR",  (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("GRID",       (0, 0), (-1, -1), 0.5, HexColor("#e5e7eb")),
            ]))
            story.append(t)

        story.append(Spacer(1, 5 * mm))
        currency = invoice.get("currency", "SGD")
        story.append(Paragraph(
            f"<b>Subtotal: {currency} {float(invoice.get('subtotal', 0)):,.2f}</b>",
            styles["Normal"],
        ))
        story.append(Paragraph(
            f"Tax: {currency} {float(invoice.get('tax_amount', 0)):,.2f}",
            styles["Normal"],
        ))
        story.append(Paragraph(
            f"<b>TOTAL: {currency} {float(invoice.get('total_amount', 0)):,.2f}</b>",
            styles["Heading2"],
        ))

        doc.build(story)
        return buffer.getvalue()

    except ImportError:
        logger.warning("reportlab not available — returning JSON bytes for invoice")
        import json
        return json.dumps(invoice).encode()


# ── AP Aging XLSX (reuses AR structure) ──────────────────────────────────────

async def generate_ap_aging_xlsx(data: dict, entity: dict) -> bytes:
    """AP Aging: same column structure as AR aging but for vendor bills."""
    # AP aging has the same structure; reuse the AR generator
    return await generate_ar_aging_xlsx(data, entity)


# ── Stub implementations for remaining report types ───────────────────────────

async def generate_balance_sheet_pdf(data: dict, entity: dict, as_at: str) -> bytes:
    """Balance Sheet PDF — stub (full implementation in a future phase)."""
    import json
    return json.dumps({"type": "balance_sheet", "entity": entity.get("name"), "as_at": as_at, "data": data}).encode()


async def generate_cash_flow_pdf(data: dict, entity: dict, period: str) -> bytes:
    """Cash Flow Statement PDF — stub (full implementation in a future phase)."""
    import json
    return json.dumps({"type": "cash_flow", "entity": entity.get("name"), "period": period, "data": data}).encode()


async def generate_gst_f5_pdf(data: dict, entity: dict, period: str) -> bytes:
    """GST F5 Return PDF — stub (full implementation in a future phase)."""
    import json
    return json.dumps({"type": "gst_f5", "entity": entity.get("name"), "period": period, "data": data}).encode()


async def generate_audit_report_pdf(data: dict, entity: dict, period: str) -> bytes:
    """Audit Trail Report PDF — stub (full implementation in a future phase)."""
    import json
    return json.dumps({"type": "audit", "entity": entity.get("name"), "period": period, "data": data}).encode()


async def generate_finance_analysis_pdf(data: dict, entity: dict) -> bytes:
    """Financial Analysis (ratios, KPIs) PDF — stub."""
    import json
    return json.dumps({"type": "analysis", "entity": entity.get("name"), "data": data}).encode()


async def generate_consolidated_report_pdf(data: dict, holding: dict) -> bytes:
    """Group Consolidated Report PDF — stub."""
    import json
    return json.dumps({"type": "consolidated", "holding": holding.get("name"), "data": data}).encode()


async def generate_quote_pdf(quote: dict, entity: dict) -> bytes:
    """Sales Quote PDF — stub (shares structure with invoice)."""
    import json
    return json.dumps({"type": "quote", "entity": entity.get("name"), "data": quote}).encode()
